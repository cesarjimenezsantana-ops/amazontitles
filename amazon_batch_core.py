"""
Core Excel processing for the Focus Camera Amazon batch updater.

This module keeps the file transformation independent from any interface:
the original command-line script and the visual web app both use these
functions.
"""

from __future__ import annotations

import os
import posixpath
import re
import unicodedata
import zipfile
from html import unescape
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import openpyxl


REFERENCE_FILENAME = "Amazon Title reference.xlsx"
ITEM_HIGHLIGHT_MAX_LENGTH = 125
TITLE_MAX_LENGTH = 200
ITEM_NAME_MAX_LENGTH = 75
TEXT_FIELD_LIMITS = {
    "bullet_point": 500,
    "special_feature": 500,
    "generic_keyword": 500,
    "model_name": 120,
    "product_description": 2000,
}

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XML_NS = "http://www.w3.org/XML/1998/namespace"

ET.register_namespace("", NS)
ET.register_namespace("r", REL_NS)
ET.register_namespace("xml", XML_NS)

BULLET_CHARACTERS = set(
    "\u2022\u2023\u25e6\u2043\u2219\u25cf\u25cb\u25aa\u25ab\u00b7"
)
BULLET_CHARACTER_RE = re.compile("[" + re.escape("".join(BULLET_CHARACTERS)) + "]")
SUSPICIOUS_ASCII_RE = re.compile(r"[{}\[\]<>\\|~^`*_#=@]")
TRADEMARK_RE = re.compile(r"[\u00ae\u00a9\u2122\u2120]")
CONTROL_CHARACTER_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
WHITESPACE_RE = re.compile(r"\s+")
SPACE_BEFORE_PUNCTUATION_RE = re.compile(r"\s+([,.;:])")
DISALLOWED_ASCII_RE = re.compile(r"[{}\[\]<>\\|^`*_#=@]")
HTML_TAG_RE = re.compile(r"<[a-zA-Z/][^>]*>")
MULTIPLE_COMMAS_RE = re.compile(r"(,\s*){2,}")
SYMBOL_REPLACEMENTS = {
    "\u20ac": "EUR",
    "\u00a3": "GBP",
    "\u00a5": "JPY",
    "\u00a2": "cents",
    "\u00d7": "x",
    "\u00f7": "/",
    "\u00b1": "+/-",
    "\u00b2": "2",
    "\u00b3": "3",
    "\u00b9": "1",
    "\u00bc": "1/4",
    "\u00bd": "1/2",
    "\u00be": "3/4",
    "\u2248": " about ",
    "\u2243": " about ",
    "\u2245": " about ",
    "\u2264": " up to ",
    "\u2265": " at least ",
    "\u00b0": " degrees ",
    "\u2013": "-",
    "\u2014": "-",
    "\u2212": "-",
    "\u2010": "-",
    "\u2011": "-",
    "\u2018": "'",
    "\u2019": "'",
    "\u201c": '"',
    "\u201d": '"',
    "\u2026": "...",
    "\u00b5": "u",
    "\u0152": "OE",
    "\u0153": "oe",
    "\u00c6": "AE",
    "\u00e6": "ae",
    "\u00df": "ss",
    "\u00d8": "O",
    "\u00f8": "o",
    "\u00a0": " ",
    "\u202f": " ",
    "\u2009": " ",
    "~": "-",
}
PROHIBITED_MARKETING_PHRASES = (
    "lifetime warranty",
    "money back guarantee",
    "money-back guarantee",
    "best seller",
    "lowest price",
    "cheapest",
    "free shipping",
    "satisfaction guaranteed",
    "#1 rated",
    "number one rated",
)
DEFAULT_TEMPLATE_COLUMNS = {
    "sku": ["C"],
    "title": ["B"],
    "item_name": ["I"],
    "item_highlight": ["J"],
    "bullet_point": [],
    "special_feature": [],
    "generic_keyword": [],
    "model_name": [],
    "product_description": [],
}
FIELD_LABELS = {
    "title": "Title",
    "item_name": "Item Name",
    "item_highlight": "Item Highlight",
    "bullet_point": "Bullet Point",
    "special_feature": "Special Feature",
    "generic_keyword": "Generic Keyword",
    "model_name": "Model Name",
    "product_description": "Product Description",
}


def q(tag: str) -> str:
    return f"{{{NS}}}{tag}"


def col_letters(ref_attr: str) -> str:
    match = re.match(r"([A-Z]+)", ref_attr or "")
    if not match:
        raise ValueError(f"Invalid cell reference: {ref_attr!r}")
    return match.group(1)


def col_num(letters: str) -> int:
    number = 0
    for char in letters:
        number = number * 26 + (ord(char) - 64)
    return number


def _clean_cell(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _format_characters(characters: list[str]) -> str:
    visible = []
    for char in characters[:10]:
        if char == "\n":
            visible.append(r"\n")
        elif char == "\r":
            visible.append(r"\r")
        elif char == "\t":
            visible.append(r"\t")
        else:
            visible.append(f"{char} (U+{ord(char):04X})")
    if len(characters) > 10:
        visible.append(f"+{len(characters) - 10} more")
    return ", ".join(visible)


def validate_bullet_point(value: str, max_length: int = ITEM_HIGHLIGHT_MAX_LENGTH) -> list[str]:
    """Return Amazon-style Bullet Point warnings for one Item Highlight value."""
    issues: list[str] = []
    text = str(value)

    if len(text) > max_length:
        issues.append(f"exceeds {max_length} characters ({len(text)})")

    if "\n" in text or "\r" in text:
        issues.append("contains line breaks")

    if "\t" in text:
        issues.append("contains tabs")

    if BULLET_CHARACTER_RE.search(text):
        issues.append("contains an actual bullet character")

    control_chars = sorted(
        {
            char
            for char in text
            if (ord(char) < 32 or ord(char) == 127) and char not in "\n\r\t"
        }
    )
    if control_chars:
        issues.append(
            "contains control characters: " + _format_characters(control_chars)
        )

    non_ascii_chars = sorted(
        {char for char in text if ord(char) > 127 and char not in BULLET_CHARACTERS}
    )
    if non_ascii_chars:
        issues.append(
            "contains non-ASCII characters: " + _format_characters(non_ascii_chars)
        )

    suspicious_ascii = sorted(set(SUSPICIOUS_ASCII_RE.findall(text)))
    if suspicious_ascii:
        issues.append(
            "contains symbols Amazon may reject: "
            + _format_characters(suspicious_ascii)
        )

    return issues


def _find_prohibited_phrases(value: str) -> list[str]:
    text = str(value or "").lower()
    return [phrase for phrase in PROHIBITED_MARKETING_PHRASES if phrase in text]


def validate_text_field(
    value: str,
    max_length: int,
    field_name: str,
    check_html: bool = False,
    check_prohibited_phrases: bool = False,
) -> list[str]:
    issues = validate_bullet_point(value, max_length=max_length)
    if check_html and HTML_TAG_RE.search(str(value)):
        issues.append(f"{field_name} contains HTML")
    if check_prohibited_phrases:
        for phrase in _find_prohibited_phrases(value):
            issues.append(f"{field_name} contains prohibited phrase: {phrase}")
    return issues


def has_unsupported_text_characters(value: str) -> bool:
    text = str(value or "")
    return bool(
        CONTROL_CHARACTER_RE.search(text)
        or BULLET_CHARACTER_RE.search(text)
        or SUSPICIOUS_ASCII_RE.search(text)
        or any(ord(char) > 127 for char in text)
    )


def _truncate_at_word(value: str, max_length: int = 500) -> str:
    if len(value) <= max_length:
        return value

    shortened = value[:max_length].rstrip(" ,;:-/")
    last_space = shortened.rfind(" ")
    if last_space >= max_length - 60:
        shortened = shortened[:last_space]
    return shortened.rstrip(" ,;:-/")


def _remove_prohibited_phrases(value: str) -> tuple[str, list[str]]:
    text = value
    actions: list[str] = []
    for phrase in PROHIBITED_MARKETING_PHRASES:
        pattern = re.compile(r"\b" + re.escape(phrase).replace(r"\ ", r"\s+") + r"\b", re.I)
        if pattern.search(text):
            text = pattern.sub("", text)
            actions.append(f"removed prohibited phrase: {phrase}")
    text = MULTIPLE_COMMAS_RE.sub(", ", text)
    text = re.sub(r"\s+([,;])", r"\1", text)
    text = re.sub(r"([,;])\s*([,;])+", r"\1", text)
    return text.strip(" ,;-"), actions


def sanitize_bullet_point(
    value: str,
    max_length: int = 500,
    remove_prohibited_phrases: bool = False,
) -> tuple[str, list[str]]:
    """Create an Amazon-friendlier Bullet Point string and list the actions."""
    original = str(value)
    text = original
    actions: list[str] = []

    if "\n" in text or "\r" in text or "\t" in text:
        text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
        actions.append("replaced line breaks or tabs with spaces")

    if CONTROL_CHARACTER_RE.search(text):
        text = CONTROL_CHARACTER_RE.sub(" ", text)
        actions.append("removed control characters")

    if BULLET_CHARACTER_RE.search(text):
        text = BULLET_CHARACTER_RE.sub(" ", text)
        actions.append("replaced bullet characters with spaces")

    if TRADEMARK_RE.search(text):
        text = TRADEMARK_RE.sub("", text)
        actions.append("removed registered, copyright, or trademark symbols")

    for source, replacement in SYMBOL_REPLACEMENTS.items():
        if source in text:
            text = text.replace(source, replacement)
            actions.append(f"replaced {source} with {replacement.strip() or 'empty'}")

    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    if ascii_text != text:
        text = ascii_text
        actions.append("converted non-ASCII characters to compatible text")

    if DISALLOWED_ASCII_RE.search(text):
        text = DISALLOWED_ASCII_RE.sub(" ", text)
        actions.append("removed disallowed ASCII symbols")

    if WHITESPACE_RE.search(text):
        collapsed = WHITESPACE_RE.sub(" ", text).strip()
        if collapsed != text:
            text = collapsed
            actions.append("normalized spaces")
    else:
        text = text.strip()

    cleaner_punctuation = SPACE_BEFORE_PUNCTUATION_RE.sub(r"\1", text)
    if cleaner_punctuation != text:
        text = cleaner_punctuation
        actions.append("fixed spaces before punctuation")

    if remove_prohibited_phrases:
        text_without_phrases, phrase_actions = _remove_prohibited_phrases(text)
        if text_without_phrases != text:
            text = text_without_phrases
            actions.extend(phrase_actions)

    shortened = _truncate_at_word(text, max_length=max_length)
    if shortened != text:
        text = shortened
        actions.append(f"trimmed text to {max_length} characters")

    if text != original and not actions:
        actions.append("normalized text")

    return text, actions


def sanitize_text_field(
    value: str,
    max_length: int,
    field_name: str,
    strip_html: bool = False,
    remove_prohibited_phrases: bool = False,
) -> tuple[str, list[str]]:
    text = str(value)
    actions: list[str] = []
    if strip_html and HTML_TAG_RE.search(text):
        text = HTML_TAG_RE.sub(" ", text)
        text = unescape(text)
        actions.append(f"removed HTML from {field_name}")
    fixed, text_actions = sanitize_bullet_point(
        text,
        max_length=max_length,
        remove_prohibited_phrases=remove_prohibited_phrases,
    )
    actions.extend(text_actions)
    return fixed, actions


def find_default_reference(base_dir: Path) -> Path | None:
    """Find the bundled reference workbook in the old or simplified layout."""
    candidates = [
        base_dir / "reference" / REFERENCE_FILENAME,
        base_dir / REFERENCE_FILENAME,
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return None


def load_reference(path: str | Path) -> dict[str, tuple[str | None, str | None]]:
    """Load the SKU reference workbook.

    Returns:
        {SKU: (item_name, item_highlight)}
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row_number = None
        header_indexes: dict[str, int] = {}

        max_scan_row = min(sheet.max_row or 20, 20)
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=max_scan_row, values_only=True),
            start=1,
        ):
            normalized = [_normalize_header(value) for value in row]
            if "sku" not in normalized:
                continue

            maybe_indexes = {name: idx for idx, name in enumerate(normalized)}
            if "item name" in maybe_indexes and "item highlight" in maybe_indexes:
                header_row_number = row_number
                header_indexes = maybe_indexes
                break

        if header_row_number is None:
            raise ValueError(
                "The reference workbook must include SKU, Item Name, and Item Highlight columns."
            )

        sku_idx = header_indexes["sku"]
        item_name_idx = header_indexes["item name"]
        item_highlight_idx = header_indexes["item highlight"]
        ref: dict[str, tuple[str | None, str | None]] = {}

        for row in sheet.iter_rows(min_row=header_row_number + 1, values_only=True):
            sku = _clean_cell(row[sku_idx] if sku_idx < len(row) else None)
            if not sku:
                continue
            item_name = _clean_cell(row[item_name_idx] if item_name_idx < len(row) else None)
            item_highlight = _clean_cell(
                row[item_highlight_idx] if item_highlight_idx < len(row) else None
            )
            ref[sku] = (item_name, item_highlight)

        return ref
    finally:
        workbook.close()


def detect_template_sheet(zin: zipfile.ZipFile) -> str | None:
    """Return the internal path for the workbook sheet named Template."""
    workbook_root = ET.fromstring(zin.read("xl/workbook.xml"))
    rels_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))

    sheet_id = None
    for sheet in workbook_root.findall(f".//{q('sheet')}"):
        if sheet.get("name") == "Template":
            sheet_id = sheet.get(f"{{{REL_NS}}}id")
            break

    if not sheet_id:
        return None

    target = None
    for rel in rels_root:
        if rel.get("Id") == sheet_id:
            target = rel.get("Target")
            break

    if not target:
        return None

    target = target.lstrip("/")
    if not target.startswith("xl/"):
        target = posixpath.normpath(posixpath.join("xl", target))
    return target


def _read_shared_strings(
    zin: zipfile.ZipFile,
) -> tuple[ET.Element | None, list[str], dict[str, int], bool]:
    try:
        root = ET.fromstring(zin.read("xl/sharedStrings.xml"))
    except KeyError:
        return None, [], {}, False

    strings: list[str] = []
    for si in root:
        text_parts = si.findall(".//" + q("t"))
        strings.append("".join(part.text or "" for part in text_parts))

    index: dict[str, int] = {}
    for idx, value in enumerate(strings):
        if value not in index:
            index[value] = idx

    return root, strings, index, True


def _append_column(columns: dict[str, list[str]], key: str, letter: str) -> None:
    if letter not in columns[key]:
        columns[key].append(letter)


def map_template_columns(sheet_data: ET.Element, cell_value) -> dict[str, list[str]]:
    columns = {key: list(value) for key, value in DEFAULT_TEMPLATE_COLUMNS.items()}
    detected = {key: [] for key in columns}

    for row in sheet_data:
        if row.tag != q("row") or int(row.get("r", "0")) != 5:
            continue
        for cell in row:
            if cell.tag != q("c") or not cell.get("r"):
                continue
            raw_value = cell_value(cell)
            if not raw_value:
                continue
            letter = col_letters(cell.get("r", ""))
            machine_name = str(raw_value).strip().lower()

            if "title_differentiation" in machine_name or "item_highlight" in machine_name:
                _append_column(detected, "item_highlight", letter)
            elif "bullet_point" in machine_name:
                _append_column(detected, "bullet_point", letter)
            elif "special_feature" in machine_name:
                _append_column(detected, "special_feature", letter)
            elif "generic_keyword" in machine_name:
                _append_column(detected, "generic_keyword", letter)
            elif "model_name" in machine_name:
                _append_column(detected, "model_name", letter)
            elif "product_description" in machine_name or "item_description" in machine_name:
                _append_column(detected, "product_description", letter)
            elif machine_name.split("[", 1)[0].strip() == "item_name":
                _append_column(detected, "item_name", letter)
            elif machine_name.startswith("::title") or (
                machine_name.startswith("title")
                and "title_differentiation" not in machine_name
                and "item_highlight" not in machine_name
            ):
                _append_column(detected, "title", letter)
            elif (
                machine_name in {"item_sku", "seller_sku", "sku"}
                or "item_sku" in machine_name
                or "seller_sku" in machine_name
            ):
                detected["sku"] = [letter]
        break

    for key, value in detected.items():
        if value:
            columns[key] = value
    return columns


def detect_data_start_row(sheet_data: ET.Element, cell_value) -> int:
    for row in sheet_data:
        if row.tag != q("row") or int(row.get("r", "0")) != 1:
            continue
        settings = " ".join(
            str(cell_value(cell) or "")
            for cell in row
            if cell.tag == q("c")
        )
        match = re.search(r"dataRow=(\d+)", settings)
        if match:
            return int(match.group(1))
        break
    return 7


def process_xlsm(
    src: str | Path,
    dst: str | Path,
    ref: dict[str, tuple[str | None, str | None]],
    auto_fix_bullets: bool = False,
    item_highlight_max_length: int = ITEM_HIGHLIGHT_MAX_LENGTH,
) -> dict[str, Any]:
    """Process one .xlsm file while preserving workbook internals."""
    with zipfile.ZipFile(src, "r") as zin:
        sheet_part = detect_template_sheet(zin)
        if sheet_part is None:
            return {"error": "The workbook does not contain a 'Template' sheet"}

        shared_root, strings, string_index, has_shared_strings = _read_shared_strings(zin)
        added_count = 0

        def idx_for(value: str) -> int:
            nonlocal added_count
            if value in string_index:
                return string_index[value]
            if shared_root is None:
                raise ValueError("The workbook does not contain sharedStrings.xml")
            si_el = ET.SubElement(shared_root, q("si"))
            t_el = ET.SubElement(si_el, q("t"))
            if value != value.strip() or "  " in value:
                t_el.set(f"{{{XML_NS}}}space", "preserve")
            t_el.text = value
            strings.append(value)
            string_index[value] = len(strings) - 1
            added_count += 1
            return string_index[value]

        def cell_val(cell: ET.Element | None) -> str | None:
            if cell is None:
                return None
            cell_type = cell.get("t")
            if cell_type == "inlineStr":
                text_el = cell.find(f"{q('is')}/{q('t')}")
                return text_el.text if text_el is not None else None

            value_el = cell.find(q("v"))
            if value_el is None or value_el.text is None:
                return None
            if cell_type == "s":
                try:
                    return strings[int(value_el.text)]
                except (IndexError, ValueError):
                    return None
            return value_el.text

        sheet_root = ET.fromstring(zin.read(sheet_part))
        sheet_data = sheet_root.find(q("sheetData"))
        if sheet_data is None:
            return {"error": "The Template sheet does not contain sheetData"}

        col_map = map_template_columns(sheet_data, cell_val)
        data_start_row = detect_data_start_row(sheet_data, cell_val)
        sku_col = (col_map.get("sku") or ["C"])[0]
        header_labels: dict[str, str] = {}
        for header_row in sheet_data:
            if header_row.tag != q("row") or int(header_row.get("r", "0")) >= data_start_row:
                continue
            for header_cell in header_row:
                if header_cell.tag != q("c") or not header_cell.get("r"):
                    continue
                header_value = str(cell_val(header_cell) or "").strip()
                if not header_value:
                    continue
                letter = col_letters(header_cell.get("r", ""))
                # Later Amazon header rows contain the stable machine field name.
                header_labels[letter] = header_value
        updated = 0
        removed = 0
        html_cleaned = 0
        field_truncated = 0
        prohibited_phrases_removed = 0
        text_fields_cleaned = 0
        title_exceeded_75 = 0
        highlight_exceeded_125 = 0
        highlight_missing = 0
        title_has_special = 0
        highlight_has_special = 0
        text_warnings: list[dict[str, Any]] = []
        text_corrections: list[dict[str, Any]] = []
        sku_contexts: list[dict[str, Any]] = []

        for row in list(sheet_data):
            if row.tag != q("row"):
                continue
            row_number = int(row.get("r", "0"))
            if row_number < 6:
                continue

            cells = {
                col_letters(cell.get("r", "")): cell
                for cell in row
                if cell.tag == q("c") and cell.get("r")
            }
            raw_sku = cell_val(cells.get(sku_col))
            sku = str(raw_sku).strip() if raw_sku is not None else ""

            shifted_data_row = (
                row_number == data_start_row - 1
                and bool(sku)
                and sku not in {"ABC123", "SKU", "contribution_sku#1.value"}
            )
            if row_number < data_start_row and not shifted_data_row:
                continue

            if not sku:
                continue

            if sku not in ref:
                sheet_data.remove(row)
                removed += 1
                continue

            sku_context: dict[str, str] = {}
            for letter, cell in sorted(cells.items(), key=lambda item: col_num(item[0])):
                raw_value = cell_val(cell)
                value = str(raw_value or "").strip()
                if not value:
                    continue
                label = header_labels.get(letter, letter)
                # Retain the column so repeated Amazon field names remain distinct.
                sku_context[f"{label} [{letter}]"] = value
            source_internal_title = next(
                (
                    str(cell_val(cells.get(letter)) or "").strip()
                    for letter in col_map.get("title", ["B"])
                    if str(cell_val(cells.get(letter)) or "").strip()
                ),
                "",
            )
            source_item_name = next(
                (
                    str(cell_val(cells.get(letter)) or "").strip()
                    for letter in col_map.get("item_name", ["I"])
                    if str(cell_val(cells.get(letter)) or "").strip()
                ),
                "",
            )
            source_highlight = next(
                (
                    str(cell_val(cells.get(letter)) or "").strip()
                    for letter in col_map.get("item_highlight", ["J"])
                    if str(cell_val(cells.get(letter)) or "").strip()
                ),
                "",
            )
            source_title_matches_item_name = bool(
                source_internal_title
                and source_item_name
                and WHITESPACE_RE.sub(" ", source_internal_title).strip().casefold()
                == WHITESPACE_RE.sub(" ", source_item_name).strip().casefold()
            )
            sku_context["Source Internal Title"] = source_internal_title
            sku_context["Source Amazon Title (Item Name)"] = source_item_name
            sku_context["Source Item Highlight"] = source_highlight
            if source_title_matches_item_name:
                sku_context["Source Internal Title Warning"] = (
                    "The uploaded Title matches Item Name. This may be a previously processed file; "
                    "verify it against the original flat file."
                )
            item_name, item_highlight = ref[sku]
            if item_name:
                sku_context["Verified reference Amazon title"] = str(item_name).strip()
            if item_highlight:
                sku_context["Verified reference Item Highlight"] = str(item_highlight).strip()
            # Keep the row element itself: rows are compacted later after SKUs
            # missing from the reference are removed.
            sku_contexts.append({"row_ref": row, "sku": sku, "context": sku_context})

            def record_warning(
                letter: str,
                field: str,
                value: str,
                issues: list[str],
            ) -> None:
                if not issues:
                    return
                preview = value[:160]
                if len(value) > 160:
                    preview += "..."
                text_warnings.append(
                    {
                        "row_ref": row,
                        "sku": sku,
                        "field": field,
                        "column": letter,
                        "issues": issues,
                        "length": len(value),
                        "preview": preview,
                    }
                )

            def record_correction(
                letter: str,
                field: str,
                original: str,
                fixed: str,
                issues_before: list[str],
                actions: list[str],
            ) -> None:
                if fixed == original:
                    return
                text_corrections.append(
                    {
                        "row_ref": row,
                        "sku": sku,
                        "field": field,
                        "column": letter,
                        "actions": actions,
                        "issues_before": issues_before,
                        "original_length": len(original),
                        "fixed_length": len(fixed),
                        "original": original,
                        "fixed": fixed,
                    }
                )

            if source_title_matches_item_name:
                title_letter = (col_map.get("title") or ["B"])[0]
                record_warning(
                    title_letter,
                    "Title",
                    source_internal_title,
                    [
                        "Uploaded Title matches Item Name; verify that this is the original flat file"
                    ],
                )

            def set_text_cell(letter: str, value: str) -> None:
                cell = cells.get(letter)
                if cell is None:
                    cell = ET.Element(q("c"))
                    cell.set("r", f"{letter}{row_number}")
                    cell.set("s", "63")
                    target_col = col_num(letter)
                    position = len(row)
                    for idx, existing in enumerate(row):
                        if col_num(col_letters(existing.get("r", ""))) > target_col:
                            position = idx
                            break
                    row.insert(position, cell)
                    cells[letter] = cell

                for child in list(cell):
                    cell.remove(child)

                if has_shared_strings:
                    cell.set("t", "s")
                    value_el = ET.SubElement(cell, q("v"))
                    value_el.text = str(idx_for(value))
                else:
                    cell.set("t", "inlineStr")
                    inline_el = ET.SubElement(cell, q("is"))
                    text_el = ET.SubElement(inline_el, q("t"))
                    if value != value.strip() or "  " in value:
                        text_el.set(f"{{{XML_NS}}}space", "preserve")
                    text_el.text = value

            def record_reference_change(letter: str, field: str, final_value: str) -> None:
                """Make reference-driven title/highlight updates reviewable too."""
                if not auto_fix_bullets:
                    return
                current_value = str(cell_val(cells.get(letter)) or "")
                if current_value == final_value:
                    return
                for correction in reversed(text_corrections):
                    if correction["row_ref"] is row and correction["column"] == letter:
                        correction["original"] = current_value
                        correction["original_length"] = len(current_value)
                        correction["actions"] = ["updated from SKU reference", *correction["actions"]]
                        return
                record_correction(
                    letter,
                    field,
                    current_value,
                    final_value,
                    [],
                    ["updated from SKU reference"],
                )

            original_titles = [
                str(cell_val(cells.get(letter)) or "").strip()
                for letter in col_map.get("title", ["B"])
            ]
            if any(len(title) > 75 for title in original_titles if title):
                title_exceeded_75 += 1
            if any(has_unsupported_text_characters(title) for title in original_titles if title):
                title_has_special += 1

            original_highlights = [
                str(cell_val(cells.get(letter)) or "").strip()
                for letter in col_map.get("item_highlight", ["J"])
            ]
            populated_highlights = [value for value in original_highlights if value]
            if not populated_highlights:
                highlight_missing += 1
            else:
                if any(len(value) > item_highlight_max_length for value in populated_highlights):
                    highlight_exceeded_125 += 1
                if any(has_unsupported_text_characters(value) for value in populated_highlights):
                    highlight_has_special += 1

            def prepare_value(
                letter: str,
                field: str,
                value: str,
                max_length: int,
                strip_html: bool = False,
                remove_prohibited: bool = False,
            ) -> str:
                nonlocal html_cleaned
                nonlocal field_truncated
                nonlocal prohibited_phrases_removed
                nonlocal text_fields_cleaned

                issues_before = validate_text_field(
                    value,
                    max_length=max_length,
                    field_name=field,
                    check_html=strip_html,
                    check_prohibited_phrases=remove_prohibited,
                )
                if auto_fix_bullets and issues_before:
                    fixed, actions = sanitize_text_field(
                        value,
                        max_length=max_length,
                        field_name=field,
                        strip_html=strip_html,
                        remove_prohibited_phrases=remove_prohibited,
                    )
                    if fixed != value:
                        record_correction(letter, field, value, fixed, issues_before, actions)
                        text_fields_cleaned += 1
                        if any(action.startswith("removed HTML") for action in actions):
                            html_cleaned += 1
                        if any(action.startswith("trimmed text") for action in actions):
                            field_truncated += 1
                        if any(action.startswith("removed prohibited phrase") for action in actions):
                            prohibited_phrases_removed += 1
                    value = fixed

                remaining_issues = validate_text_field(
                    value,
                    max_length=max_length,
                    field_name=field,
                    check_html=strip_html,
                    check_prohibited_phrases=remove_prohibited,
                )
                record_warning(letter, field, value, remaining_issues)
                return value

            # Title is a read-only source field from the uploaded Amazon file.
            # Preserve it byte-for-byte; reference data only updates Item Name.

            if item_name:
                item_name_value = prepare_value(
                    (col_map.get("item_name") or ["I"])[0],
                    "Item Name",
                    item_name,
                    ITEM_NAME_MAX_LENGTH,
                )
                for letter in col_map.get("item_name", ["I"]):
                    record_reference_change(letter, "Item Name", item_name_value)
                    set_text_cell(letter, item_name_value)
            if item_highlight:
                original_highlight = str(item_highlight)
                fixed_highlight = prepare_value(
                    (col_map.get("item_highlight") or ["J"])[0],
                    "Item Highlight",
                    original_highlight,
                    max_length=item_highlight_max_length,
                    remove_prohibited=True,
                )
                for letter in col_map.get("item_highlight", ["J"]):
                    record_reference_change(letter, "Item Highlight", fixed_highlight)
                    set_text_cell(letter, fixed_highlight)

            for field_key, max_length in TEXT_FIELD_LIMITS.items():
                field = FIELD_LABELS[field_key]
                for letter in col_map.get(field_key, []):
                    value = cell_val(cells.get(letter))
                    if not value:
                        continue
                    original_value = str(value)
                    fixed_value = prepare_value(
                        letter,
                        field,
                        original_value,
                        max_length=max_length,
                        strip_html=field_key == "product_description",
                    )
                    if fixed_value != original_value:
                        set_text_cell(letter, fixed_value)

            updated += 1

        header_rows: list[ET.Element] = []
        data_rows: list[ET.Element] = []
        for row in list(sheet_data):
            if row.tag != q("row"):
                continue
            row_number = int(row.get("r", "0"))
            cells = {
                col_letters(cell.get("r", "")): cell
                for cell in row
                if cell.tag == q("c") and cell.get("r")
            }
            raw_sku = cell_val(cells.get(sku_col))
            sku = str(raw_sku).strip() if raw_sku is not None else ""
            shifted_data_row = (
                row_number == data_start_row - 1
                and bool(sku)
                and sku not in {"ABC123", "SKU", "contribution_sku#1.value"}
            )
            if row_number < data_start_row and not shifted_data_row:
                header_rows.append(row)
            else:
                data_rows.append(row)

        for child in list(sheet_data):
            sheet_data.remove(child)

        for row in sorted(header_rows, key=lambda item: int(item.get("r", "0"))):
            sheet_data.append(row)

        def renumber_row(row: ET.Element, new_number: int) -> None:
            row.set("r", str(new_number))
            for cell in row:
                if cell.tag == q("c") and cell.get("r"):
                    letters = col_letters(cell.get("r", ""))
                    cell.set("r", f"{letters}{new_number}")

        for index, row in enumerate(data_rows, start=data_start_row):
            renumber_row(row, index)
            sheet_data.append(row)

        if data_rows:
            final_rows = data_start_row + len(data_rows) - 1
        elif header_rows:
            final_rows = max(int(row.get("r", "0")) for row in header_rows)
        else:
            final_rows = 1

        dimension = sheet_root.find(q("dimension"))
        if dimension is not None:
            dimension.set("ref", re.sub(r"\d+$", str(final_rows), dimension.get("ref", "A1:A1")))

        if shared_root is not None:
            old_count = int(shared_root.get("count", "0"))
            shared_root.set("count", str(old_count + added_count))
            shared_root.set("uniqueCount", str(len(strings)))

        warnings = []
        for warning in text_warnings:
            row_ref = warning.pop("row_ref")
            warnings.append(
                {
                    **warning,
                    "row": int(row_ref.get("r", "0")),
                }
            )

        corrections = []
        for correction in text_corrections:
            row_ref = correction.pop("row_ref")
            corrections.append(
                {
                    **correction,
                    "row": int(row_ref.get("r", "0")),
                }
            )

        resolved_sku_contexts = []
        for sku_context in sku_contexts:
            row_ref = sku_context["row_ref"]
            resolved_sku_contexts.append(
                {
                    "row": int(row_ref.get("r", "0")),
                    "sku": sku_context["sku"],
                    "context": sku_context["context"],
                }
            )

        os.makedirs(os.path.dirname(os.fspath(dst)) or ".", exist_ok=True)
        with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == sheet_part:
                    data = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
                elif item.filename == "xl/sharedStrings.xml" and shared_root is not None:
                    data = ET.tostring(shared_root, encoding="utf-8", xml_declaration=True)
                else:
                    data = zin.read(item.filename)
                zout.writestr(item, data)

    return {
        "updated": updated,
        "removed": removed,
        "new_strings": added_count,
        "final_rows": final_rows,
        "warnings": warnings,
        "sku_contexts": resolved_sku_contexts,
        "corrections": corrections,
        "html_cleaned": html_cleaned,
        "field_truncated": field_truncated,
        "prohibited_phrases_removed": prohibited_phrases_removed,
        "text_fields_cleaned": text_fields_cleaned,
        "title_exceeded_75": title_exceeded_75,
        "highlight_exceeded_125": highlight_exceeded_125,
        "highlight_missing": highlight_missing,
        "title_has_special": title_has_special,
        "highlight_has_special": highlight_has_special,
    }


def update_xlsm_cells(path: str | Path, changes: list[dict[str, Any]]) -> None:
    """Update text cells in the Template sheet without rebuilding the workbook."""
    workbook_path = Path(path)
    if not changes:
        return

    with zipfile.ZipFile(workbook_path, "r") as zin:
        sheet_part = detect_template_sheet(zin)
        if sheet_part is None:
            raise ValueError("The workbook does not contain a 'Template' sheet")
        shared_root, strings, string_index, has_shared_strings = _read_shared_strings(zin)
        sheet_root = ET.fromstring(zin.read(sheet_part))
        cells = {
            cell.get("r"): cell
            for cell in sheet_root.iter(q("c"))
            if cell.get("r")
        }

        def shared_index(value: str) -> int:
            if value in string_index:
                return string_index[value]
            if shared_root is None:
                raise ValueError("The workbook does not contain sharedStrings.xml")
            si_el = ET.SubElement(shared_root, q("si"))
            text_el = ET.SubElement(si_el, q("t"))
            if value != value.strip() or "  " in value:
                text_el.set(f"{{{XML_NS}}}space", "preserve")
            text_el.text = value
            strings.append(value)
            string_index[value] = len(strings) - 1
            return string_index[value]

        for change in changes:
            ref = f"{change['column']}{int(change['row'])}"
            cell = cells.get(ref)
            if cell is None:
                raise ValueError(f"Cell {ref} was not found")
            value = str(change.get("value", ""))
            for child in list(cell):
                cell.remove(child)
            if has_shared_strings:
                cell.set("t", "s")
                value_el = ET.SubElement(cell, q("v"))
                value_el.text = str(shared_index(value))
            else:
                cell.set("t", "inlineStr")
                inline_el = ET.SubElement(cell, q("is"))
                text_el = ET.SubElement(inline_el, q("t"))
                if value != value.strip() or "  " in value:
                    text_el.set(f"{{{XML_NS}}}space", "preserve")
                text_el.text = value

        if shared_root is not None:
            shared_root.set("uniqueCount", str(len(strings)))
            shared_root.set("count", str(max(int(shared_root.get("count", "0")), len(strings))))

        replacement = workbook_path.with_suffix(".reviewed.tmp")
        with zipfile.ZipFile(replacement, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == sheet_part:
                    data = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
                elif item.filename == "xl/sharedStrings.xml" and shared_root is not None:
                    data = ET.tostring(shared_root, encoding="utf-8", xml_declaration=True)
                else:
                    data = zin.read(item.filename)
                zout.writestr(item, data)
    replacement.replace(workbook_path)
