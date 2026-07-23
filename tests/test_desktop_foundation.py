from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from urllib.request import urlopen

from openpyxl import Workbook, load_workbook

from amazon_batch_core import (
    load_reference,
    process_xlsm,
    sanitize_bullet_point,
    update_xlsm_cells,
)
from app_paths import RESOURCE_DIR
from desktop_app import LocalApplicationServer


class DesktopFoundationTests(unittest.TestCase):
    def test_packaged_resources_exist(self) -> None:
        self.assertTrue((RESOURCE_DIR / "templates" / "index.html").is_file())
        self.assertTrue((RESOURCE_DIR / "static" / "app.js").is_file())
        self.assertTrue(any(RESOURCE_DIR.glob("*.xlsx")))

    def test_web_assets_are_versioned_to_avoid_stale_desktop_cache(self) -> None:
        template = (RESOURCE_DIR / "templates" / "index.html").read_text(encoding="utf-8")
        self.assertIn("filename='styles.css', v=app_version", template)
        self.assertIn("filename='app.js', v=app_version", template)

    def test_range_separator_remains_readable(self) -> None:
        fixed, actions = sanitize_bullet_point("Range: 120~150 miles")
        self.assertEqual(fixed, "Range: 120-150 miles")
        self.assertIn("replaced ~ with -", actions)

    def test_reference_loader_ignores_title_column(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            reference = Path(temp_dir) / "reference.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["SKU", "Title", "Item Name", "Item Highlight"])
            sheet.append([
                "SKU-1",
                "This reference Title must be ignored",
                "New Item Name",
                "New Item Highlight",
            ])
            workbook.save(reference)

            self.assertEqual(
                load_reference(reference),
                {"SKU-1": ("New Item Name", "New Item Highlight")},
            )

    def test_reference_loader_accepts_highlights_alias(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            reference = Path(temp_dir) / "reference.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["SKU", "Title", "Item Name", "Highlights"])
            sheet.append(["SKU-1", "Ignore this", "New Item Name", "New Highlights"])
            workbook.save(reference)

            self.assertEqual(
                load_reference(reference),
                {"SKU-1": ("New Item Name", "New Highlights")},
            )

    def test_reference_fields_never_overwrite_uploaded_title(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.xlsm"
            output = Path(temp_dir) / "output.xlsm"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Template"
            sheet["A1"] = "settings=dataRow=7"
            sheet["B4"], sheet["C4"], sheet["I4"], sheet["J4"] = (
                "Title", "SKU", "Item Name", "Item Highlight"
            )
            sheet["B5"], sheet["C5"], sheet["I5"], sheet["J5"] = (
                "::title",
                "item_sku",
                "item_name[marketplace_id=US]#1.value",
                "item_highlight[marketplace_id=US]#1.value",
            )
            sheet["B7"], sheet["C7"], sheet["I7"], sheet["J7"] = (
                "Original ~ internal catalog title",
                "SKU-1",
                "Old Amazon title",
                "Old highlight",
            )
            sheet["B8"], sheet["C8"], sheet["I8"] = (
                "Already duplicated title",
                "SKU-2",
                "Already duplicated title",
            )
            workbook.save(source)

            stats = process_xlsm(
                source,
                output,
                {
                    "SKU-1": ("New reference Amazon title", "New reference highlight"),
                    "SKU-2": ("Second reference title", None),
                },
                auto_fix_bullets=True,
            )

            self.assertNotIn("error", stats)
            processed = load_workbook(output, read_only=True)["Template"]
            self.assertEqual(processed["B7"].value, "Original ~ internal catalog title")
            self.assertEqual(processed["I7"].value, "New reference Amazon title")
            self.assertEqual(processed["J7"].value, "New reference highlight")
            self.assertEqual(processed["B8"].value, "Already duplicated title")
            sku_1_context = next(
                item["context"] for item in stats["sku_contexts"] if item["sku"] == "SKU-1"
            )
            self.assertEqual(
                sku_1_context["Source Uploaded Title"],
                "Original ~ internal catalog title",
            )
            self.assertEqual(
                sku_1_context["Source Amazon Title (Item Name)"],
                "Old Amazon title",
            )
            title_warnings = [
                warning for warning in stats["warnings"] if warning["field"] == "Title"
            ]
            self.assertEqual(title_warnings, [])

    def test_shifted_columns_preserve_title_and_update_reference_fields(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.xlsm"
            output = Path(temp_dir) / "output.xlsm"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Template"
            sheet["A1"] = "settings=dataRow=7"
            sheet["D4"], sheet["F4"], sheet["K4"], sheet["M4"] = (
                "Title", "SKU", "Item Name", "Item Highlight"
            )
            sheet["D5"], sheet["F5"], sheet["K5"], sheet["M5"] = (
                "::title",
                "item_sku",
                "item_name[marketplace_id=US]#1.value",
                "item_highlight[marketplace_id=US]#1.value",
            )
            sheet["D7"], sheet["F7"], sheet["K7"], sheet["M7"] = (
                "Original shifted Title",
                "SKU-1",
                "Old Item Name",
                "Old Highlight",
            )
            workbook.save(source)

            stats = process_xlsm(
                source,
                output,
                {"SKU-1": ("Reference Item Name", "Reference Highlights")},
            )

            self.assertNotIn("error", stats)
            processed = load_workbook(output, read_only=True)["Template"]
            self.assertEqual(processed["D7"].value, "Original shifted Title")
            self.assertEqual(processed["K7"].value, "Reference Item Name")
            self.assertEqual(processed["M7"].value, "Reference Highlights")

    def test_review_updates_reject_title_even_when_shifted(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "review.xlsm"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Template"
            sheet["D5"] = "::title"
            sheet["K5"] = "item_name[marketplace_id=US]#1.value"
            sheet["D7"] = "Original Title"
            sheet["K7"] = "Original Item Name"
            workbook.save(workbook_path)

            with self.assertRaisesRegex(ValueError, "read-only"):
                update_xlsm_cells(
                    workbook_path,
                    [{"column": "D", "row": 7, "value": "Forbidden Title"}],
                )

            update_xlsm_cells(
                workbook_path,
                [{"column": "K", "row": 7, "value": "Reviewed Item Name"}],
            )
            processed = load_workbook(workbook_path, read_only=True)["Template"]
            self.assertEqual(processed["D7"].value, "Original Title")
            self.assertEqual(processed["K7"].value, "Reviewed Item Name")

    def test_local_server_uses_loopback_and_random_port(self) -> None:
        server = LocalApplicationServer()
        server.start()
        try:
            self.assertRegex(server.url, r"^http://127\.0\.0\.1:\d+/$")
            with urlopen(f"{server.url}api/health", timeout=5) as response:
                payload = json.load(response)
            self.assertEqual(payload, {"ok": True})
        finally:
            server.stop()


if __name__ == "__main__":
    unittest.main()
